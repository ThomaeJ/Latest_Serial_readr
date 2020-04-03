import glob
import pandas as pd
import regex as re
from collections import defaultdict
from openpyxl import Workbook
from openpyxl import load_workbook
import os

qual = input("Quality file name:")
temp = input("Template file name:")

inputs = {
    'qual' : qual,
    'temp' : temp
}

#glob
path = r"C:\Serial_Reader-master\Serial_Reader\Quality\\" +(inputs.get('qual'))
#path = r"C:\Users\thoma\Documents\WGU\Misc\Serial_Reader\Serial_Reader\Quality\SAL2342004V_Quality.log"
path2 = r"Templates\\" +(inputs.get('temp'))
#path2 = r"C:\Users\thoma\Documents\WGU\Misc\Serial_Reader\Serial_Reader\Templates\WIP1.xlsx"

regex = ['"0/RP0"', '"0/RP1"', '"0/FC0"', '"0/FC1"', '"0/FC2"', '"0/FC3"', '"0/CI0"',
         '"0/EC0"', '"ECU-DISK-SSD0"', '"ECU-DISK-SSD1"', '"NCS 4009 shelf assembly - DC Power"',
         '"0/FT0"', '"0/FT2"', '"0/FT3"', '"0/FT4"', '"0/FT5"', '"0/PT0"', '"0/PT0-PM0"', '"0/PT0-PM1"',
         '"0/PT0-PM2"', '"0/PT0-PM3"', '"0/PT1"', '"0/PT1-PM0"', '"0/PT1-PM1"', '"0/PT1-PM2"', '"0/PT1-PM3"']

part_list = []
serial_list = []

def get_serials(p):
    s = p.split()
    scrum = s.index('PID:')
    bum = s.index('SN:')
    part_list.append(s[scrum+1])
    serial_list.append(s[bum+1])

with open(path, "r") as quality:
    file = quality.readlines()
    files = iter(file)
    for line in files:
        for r in regex:
            if r in line:
                p = next(files)
                get_serials(p)
                print(p)


print (part_list, serial_list)

print(os.getcwd())
print(os.listdir())

template = load_workbook(path2)
sheets = template.sheetnames
server = template[sheets[1]]
cell_range = server['I1' : 'I100']

for num in range(1, 50):
    i = server['I' + str(num)]
    j = server['J' + str(num)]
    if i.value in serial_list:
        j.value = i.value

template.save(inputs.get('temp'))


#print(template.active)
#d = server['I1']
#print(d.value)


    #d = server[i]
    #print (d.value)

        #print(f)

#template = load_workbook(filename=path2, data_only=True)
#server = template['SER-VER']
#print (server['I50'])


#template = pd.read_excel(open(path2, 'rb'), sheet_name= 'SER-VER')



